#!/usr/bin/env python
"""Excel de evaluación humana COMPARATIVA: misma interacción, 4 versiones ciegas.

- 16 interacciones del subset de 50, estratificadas por product_category
  (≥2 por categoría), seed 42, restringidas a las cubiertas por LOS 4
  generadores (langgraph, crewai, openai_agents, baseline_prompt) en run_1.
- Por interacción se muestran los 4 artículos como Versión A/B/C/D en orden
  aleatorio (seed 42) → cegado. Total 16×4 = 64 artículos.

Hojas: `evaluacion` (agrupada por interacción), `key` (id×versión→framework),
`rubrica` (misma de build_human_eval.py).

Uso:
    .venv/bin/python scripts/build_human_eval_comparative.py
"""
from __future__ import annotations

import json
import random
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
from scripts.build_human_eval import RUBRICA, SCORE_DIMS, _fmt_resolution  # noqa: E402

SEED = 42
GENERATORS = ["langgraph", "crewai", "openai_agents", "baseline_prompt"]
N_INTERACTIONS = 16
MIN_PER_CATEGORY = 2
VERSIONS = ["A", "B", "C", "D"]
OUT = PROJECT_ROOT / "eval" / "rubrics" / "human_evaluation_comparative.xlsx"


# ---------------------------------------------------------------------------
# Carga
# ---------------------------------------------------------------------------
def load_interactions() -> Dict[str, Dict[str, Any]]:
    out = {}
    for line in (PROJECT_ROOT / "data/processed/interactions.jsonl").read_text(encoding="utf-8").splitlines():
        if line.strip():
            d = json.loads(line)
            out[d["interaction_id"]] = d
    return out


def interaction_topic(inter: Dict[str, Any]) -> str:
    """Tema = primer turno del cliente (la pregunta), recortado."""
    for t in inter.get("turns", []):
        if t.get("role") == "cliente":
            msg = (t.get("message") or "").strip()
            return msg[:160] + ("…" if len(msg) > 160 else "")
    md = inter.get("metadata", {})
    return f"{md.get('product_specific','')} — {md.get('gap_topic','')}".strip(" —")


def framework_interaction_articles(fw: str) -> Dict[str, Dict[str, Any]]:
    """interaction_id → registro de artículo (run_1) que la cubre (primero)."""
    rd = PROJECT_ROOT / "runs/experiment" / fw / "run_1"
    arts = {
        a["article_id"]: a
        for a in (json.loads(l) for l in (rd / "generated_articles.jsonl").read_text(encoding="utf-8").splitlines() if l.strip())
    }
    amap = json.loads((rd / "article_interaction_map.json").read_text(encoding="utf-8"))
    rev: Dict[str, Dict[str, Any]] = {}
    for aid, iids in amap.items():
        iids = [iids] if isinstance(iids, str) else iids
        for i in iids:
            rev.setdefault(i, arts.get(aid))
    return rev


# ---------------------------------------------------------------------------
# Selección estratificada (≥2 por categoría, seed)
# ---------------------------------------------------------------------------
def select_interactions(candidates: List[str], inter: Dict[str, Any], rng: random.Random) -> List[str]:
    by_cat: Dict[str, List[str]] = defaultdict(list)
    for i in candidates:
        by_cat[inter[i]["metadata"]["product_category"]].append(i)
    cats = sorted(by_cat)
    for c in cats:
        rng.shuffle(by_cat[c])
    alloc = {c: MIN_PER_CATEGORY for c in cats}
    extra = N_INTERACTIONS - sum(alloc.values())
    # repartir extras a las categorías con más disponibilidad (desempate aleatorio)
    order = sorted(cats, key=lambda c: (len(by_cat[c]), rng.random()), reverse=True)
    k = 0
    while extra > 0:
        c = order[k % len(order)]
        if alloc[c] < len(by_cat[c]):
            alloc[c] += 1
            extra -= 1
        k += 1
        if k > 1000:
            break
    chosen: List[str] = []
    for c in cats:
        chosen.extend(by_cat[c][: alloc[c]])
    rng.shuffle(chosen)
    return chosen[:N_INTERACTIONS]


# ---------------------------------------------------------------------------
# Formateo
# ---------------------------------------------------------------------------
def evidence_short(body: Any) -> str:
    if not isinstance(body, dict):
        return ""
    ep = body.get("evidence_pack") or {}
    if not isinstance(ep, dict):
        return str(ep)[:300]
    iids = ep.get("interaction_ids") or []
    frags = ep.get("key_fragments") or []
    cem = ep.get("claim_evidence_map") or {}
    parts = []
    if iids:
        parts.append("Interacciones: " + ", ".join(iids))
    parts.append(f"{len(frags)} fragmento(s) de evidencia")
    if isinstance(cem, dict) and cem:
        claims = list(cem.keys())[:3]
        parts.append("Afirmaciones clave:\n  - " + "\n  - ".join(claims))
        if len(cem) > 3:
            parts.append(f"  (+{len(cem)-3} más)")
    elif frags:
        parts.append("Ej.: " + str(frags[0])[:160])
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
SEP_FILL = PatternFill("solid", fgColor="D9D9D9")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GROUP_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

COLS = ["version", "title", "resolution", "evidence_pack"] + SCORE_DIMS + ["comentarios"]
SCORE_START = 5  # 1-based index de 'claridad'


def build_eval_sheet(ws, groups: List[Dict[str, Any]]):
    # encabezado global
    ws.append([c if c not in SCORE_DIMS else f"{c} (1-5)" for c in COLS])
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"

    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                        allow_blank=True, showErrorMessage=True,
                        error="Ingresa un entero de 1 a 5", errorTitle="Valor inválido")
    ws.add_data_validation(dv)

    r = 2
    for g in groups:
        # fila de encabezado de la interacción (fusionada)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        hc = ws.cell(row=r, column=1)
        hc.value = f"{g['interaction_id']}  —  {g['topic']}"
        hc.fill = GROUP_FILL
        hc.font = GROUP_FONT
        hc.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 40
        r += 1
        # 4 versiones
        for v in g["versions"]:  # ya en orden A,B,C,D
            ws.cell(row=r, column=1, value=v["version"])
            ws.cell(row=r, column=2, value=v["title"])
            ws.cell(row=r, column=3, value=v["resolution"])
            ws.cell(row=r, column=4, value=v["evidence"])
            for col in range(1, len(COLS) + 1):
                cell = ws.cell(row=r, column=col)
                cell.alignment = WRAP_TOP
                cell.border = BORDER
                if SCORE_START <= col <= SCORE_START + len(SCORE_DIMS) - 1:
                    cell.fill = SCORE_FILL
                    dv.add(cell)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            ws.row_dimensions[r].height = 150
            r += 1
        # separador
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        ws.cell(row=r, column=1).fill = SEP_FILL
        ws.row_dimensions[r].height = 8
        r += 1

    widths = {"version": 9, "title": 34, "resolution": 56, "evidence_pack": 46, "comentarios": 28}
    for i, c in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)


def build_key_sheet(ws, groups: List[Dict[str, Any]]):
    cols = ["interaction_id", "version", "framework", "orig_article_id", "product_category"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = HEADER_FONT
    ws.freeze_panes = "A2"
    for g in groups:
        for v in g["versions"]:
            ws.append([g["interaction_id"], v["version"], v["framework"],
                       v["orig_article_id"], g["product_category"]])
    widths = [16, 9, 18, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_rubric_sheet(ws):
    cols = ["dimensión", "definición", "nivel 1", "nivel 2", "nivel 3", "nivel 4", "nivel 5", "ejemplos ancla"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = HEADER_FONT
    ws.freeze_panes = "A2"
    for dim in SCORE_DIMS:
        rb = RUBRICA[dim]
        ws.append([dim, rb["_def"], rb["1"], rb["2"], rb["3"], rb["4"], rb["5"], rb["_ancla"]])
    for i, w in enumerate([16, 40, 30, 30, 30, 30, 34, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in range(2, len(SCORE_DIMS) + 2):
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).alignment = WRAP_TOP
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 150
    ws.append([])
    ws.append(["Evaluación comparativa CIEGA: por cada interacción puntúa las 4 versiones "
               "(A-D) en cada dimensión 1-5. El orden A/B/C/D es aleatorio por interacción, "
               "así que la misma letra NO corresponde al mismo framework entre grupos."])


def main() -> int:
    rng = random.Random(SEED)
    inter = load_interactions()
    subset = set(yaml.safe_load(open(PROJECT_ROOT / "data/splits/eval_subset_50.yaml"))["interaction_ids"])
    maps = {fw: framework_interaction_articles(fw) for fw in GENERATORS}
    candidates = [i for i in subset if all(i in maps[fw] and maps[fw][i] for fw in GENERATORS)]
    print(f"Candidatas (subset ∩ cubiertas por los 4): {len(candidates)}")

    chosen = select_interactions(candidates, inter, rng)
    dist = Counter(inter[i]["metadata"]["product_category"] for i in chosen)
    print(f"Seleccionadas: {len(chosen)} → {dict(dist)}")

    groups: List[Dict[str, Any]] = []
    for iid in chosen:
        # orden aleatorio de frameworks → versiones A,B,C,D
        fw_order = GENERATORS[:]
        rng.shuffle(fw_order)
        versions = []
        for label, fw in zip(VERSIONS, fw_order):
            rec = maps[fw][iid]
            body = rec.get("article")
            title = body.get("title", "") if isinstance(body, dict) else ""
            resolution = _fmt_resolution(body.get("resolution")) if isinstance(body, dict) else str(body or "")
            versions.append({
                "version": label,
                "framework": fw,
                "orig_article_id": rec.get("article_id"),
                "title": title,
                "resolution": resolution,
                "evidence": evidence_short(body),
            })
        groups.append({
            "interaction_id": iid,
            "product_category": inter[iid]["metadata"]["product_category"],
            "topic": interaction_topic(inter[iid]),
            "versions": versions,
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "evaluacion"
    build_eval_sheet(ws, groups)
    build_key_sheet(wb.create_sheet("key"), groups)
    build_rubric_sheet(wb.create_sheet("rubrica"))
    wb.save(OUT)

    total = sum(len(g["versions"]) for g in groups)
    print(f"\n✅ {len(groups)} interacciones × 4 = {total} artículos → {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
