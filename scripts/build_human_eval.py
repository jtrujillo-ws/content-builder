#!/usr/bin/env python
"""Prepara el Excel de evaluación humana (ciega) a partir del run 1 principal.

- 15 artículos por framework (langgraph, crewai, openai_agents), estratificados
  por product_category → 45.
- 5 artículos de baseline_prompt como control → 50 total.
- Orden aleatorio (seed 42) y blind id EVAL-### (sin revelar el framework).
- 3 hojas: `evaluacion` (a puntuar), `key` (mapeo id→framework), `rubrica`.

Uso:
    .venv/bin/python scripts/build_human_eval.py
"""
from __future__ import annotations

import json
import random
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SEED = 42
FRAMEWORKS = ["langgraph", "crewai", "openai_agents"]
N_PER_FRAMEWORK = 15
CONTROL_FW = "baseline_prompt"
N_CONTROL = 5
SCORE_DIMS = ["claridad", "exactitud", "completitud", "aplicabilidad", "consistencia"]

OUT = PROJECT_ROOT / "eval" / "rubrics" / "human_evaluation.xlsx"


# ---------------------------------------------------------------------------
# Carga de datos
# ---------------------------------------------------------------------------
def load_interaction_categories() -> Dict[str, str]:
    cats: Dict[str, str] = {}
    path = PROJECT_ROOT / "data" / "processed" / "interactions.jsonl"
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        d = json.loads(line)
        cats[d["interaction_id"]] = (d.get("metadata") or {}).get("product_category") or "otros"
    return cats


def _article_interactions(article: Dict[str, Any], amap: Dict[str, Any]) -> List[str]:
    aid = article.get("article_id")
    ids = amap.get(aid)
    if isinstance(ids, str):
        ids = [ids]
    if not ids:
        # fallback: del propio evidence_pack
        body = article.get("article")
        if isinstance(body, dict):
            ep = body.get("evidence_pack") or {}
            ids = ep.get("interaction_ids") or []
    return list(ids or [])


def load_framework_articles(fw: str, cats: Dict[str, str]) -> List[Dict[str, Any]]:
    run_dir = PROJECT_ROOT / "runs" / "experiment" / fw / "run_1"
    arts = [
        json.loads(l)
        for l in (run_dir / "generated_articles.jsonl").read_text(encoding="utf-8").splitlines()
        if l.strip()
    ]
    amap_path = run_dir / "article_interaction_map.json"
    amap = json.loads(amap_path.read_text(encoding="utf-8")) if amap_path.exists() else {}
    out = []
    for a in arts:
        iids = _article_interactions(a, amap)
        # categoría = mayoría entre sus interacciones (desempate: primera)
        cat_counts = Counter(cats.get(i, "otros") for i in iids) if iids else Counter(["otros"])
        category = cat_counts.most_common(1)[0][0]
        out.append({
            "framework": fw,
            "orig_article_id": a.get("article_id"),
            "interaction_ids": iids,
            "product_category": category,
            "article": a.get("article"),
        })
    return out


# ---------------------------------------------------------------------------
# Muestreo estratificado (largest remainder, seed fijo)
# ---------------------------------------------------------------------------
def stratified_sample(items: List[Dict[str, Any]], n: int, rng: random.Random) -> List[Dict[str, Any]]:
    by_cat: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for it in items:
        by_cat[it["product_category"]].append(it)
    cats = sorted(by_cat)
    total = len(items)
    # cuota proporcional con reparto de residuos (largest remainder)
    raw = {c: n * len(by_cat[c]) / total for c in cats}
    alloc = {c: int(raw[c]) for c in cats}
    remainder = n - sum(alloc.values())
    # repartir el resto por mayor parte fraccionaria
    for c in sorted(cats, key=lambda c: (raw[c] - alloc[c]), reverse=True)[:remainder]:
        alloc[c] += 1
    # no pedir más de lo disponible por categoría; reasignar sobrante
    chosen: List[Dict[str, Any]] = []
    leftover_pool: List[Dict[str, Any]] = []
    for c in cats:
        pool = by_cat[c][:]
        rng.shuffle(pool)
        take = min(alloc[c], len(pool))
        chosen.extend(pool[:take])
        leftover_pool.extend(pool[take:])
    # completar si faltan (categorías sin suficientes)
    rng.shuffle(leftover_pool)
    while len(chosen) < n and leftover_pool:
        chosen.append(leftover_pool.pop())
    return chosen[:n]


# ---------------------------------------------------------------------------
# Formateo legible de los campos KCS
# ---------------------------------------------------------------------------
def _fmt_environment(v: Any) -> str:
    if not isinstance(v, dict):
        return str(v or "")
    parts = []
    for k, val in v.items():
        if isinstance(val, list):
            val = ", ".join(str(x) for x in val)
        parts.append(f"{k}: {val}")
    return "\n".join(parts)


def _fmt_resolution(v: Any) -> str:
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return str(v or "")


def _fmt_evidence(v: Any) -> str:
    if not isinstance(v, dict):
        return str(v or "")
    lines = []
    iids = v.get("interaction_ids")
    if iids:
        lines.append("Interacciones: " + ", ".join(iids))
    frags = v.get("key_fragments") or []
    if frags:
        lines.append("\nFragmentos de evidencia:")
        for f in frags:
            lines.append(f"  • {f}")
    cem = v.get("claim_evidence_map")
    if isinstance(cem, dict) and cem:
        lines.append("\nMapeo afirmación → evidencia:")
        for claim, ev in cem.items():
            ev_s = ", ".join(ev) if isinstance(ev, list) else str(ev)
            lines.append(f"  • {claim} ⟵ {ev_s}")
    return "\n".join(lines)


def article_fields(body: Any) -> Dict[str, str]:
    """Extrae los 7 campos KCS legibles. Robusto si `article` es str."""
    if not isinstance(body, dict):
        return {"title": "", "environment": "", "problem": str(body or ""),
                "cause": "", "resolution": "", "evidence_pack": ""}
    return {
        "title": str(body.get("title", "") or ""),
        "environment": _fmt_environment(body.get("environment")),
        "problem": str(body.get("problem", "") or ""),
        "cause": str(body.get("cause", "") or ""),
        "resolution": _fmt_resolution(body.get("resolution")),
        "evidence_pack": _fmt_evidence(body.get("evidence_pack")),
    }


# ---------------------------------------------------------------------------
# Rúbrica
# ---------------------------------------------------------------------------
RUBRICA: Dict[str, Dict[str, str]] = {
    "claridad": {
        "_def": "¿El artículo se entiende fácilmente? Redacción, organización y ausencia de ambigüedad.",
        "1": "Confuso e incomprensible; redacción desorganizada que impide entender el contenido.",
        "2": "Difícil de seguir; hay que releer varias veces para captar la idea.",
        "3": "Comprensible con esfuerzo; estructura irregular o frases ambiguas.",
        "4": "Claro y bien organizado; pocas ambigüedades.",
        "5": "Excepcionalmente claro; lenguaje preciso y estructura impecable.",
        "_ancla": "Ancla 5: títulos descriptivos, pasos numerados sin jerga; Ancla 1: párrafo único sin estructura, términos sin explicar.",
    },
    "exactitud": {
        "_def": "¿La información es correcta y fiel a la evidencia (sin invenciones/alucinaciones)?",
        "1": "Errores graves o afirmaciones inventadas sin respaldo en la interacción fuente.",
        "2": "Varias imprecisiones o afirmaciones no respaldadas por la evidencia.",
        "3": "Mayormente correcto con alguna imprecisión menor.",
        "4": "Correcto y fiel a la evidencia; sin invenciones.",
        "5": "Totalmente exacto; cada afirmación es trazable a un fragmento de la fuente.",
        "_ancla": "Ancla 5: 'servicio gratuito' ↔ INT-2024-010 turn 6; Ancla 1: inventa un costo o un canal no mencionado.",
    },
    "completitud": {
        "_def": "¿Cubre toda la información necesaria para resolver el caso?",
        "1": "Omite información esencial; el artículo es inservible.",
        "2": "Faltan pasos o datos importantes para resolver el caso.",
        "3": "Cubre lo básico pero omite detalles útiles (costos, canales alternos).",
        "4": "Completo; cubre el caso y sus variantes principales.",
        "5": "Exhaustivo; incluye canales alternativos, costos, excepciones y requisitos.",
        "_ancla": "Ancla 5: app + sucursal + costo + entrega por correo; Ancla 1: solo menciona 'use la app' sin pasos.",
    },
    "aplicabilidad": {
        "_def": "¿El usuario puede ejecutar la solución siguiendo el artículo?",
        "1": "No ofrece pasos aplicables.",
        "2": "Pasos vagos o incompletos; el usuario no sabría qué hacer.",
        "3": "Pasos seguibles pero con saltos o supuestos no explicados.",
        "4": "Pasos concretos y ordenados, fáciles de ejecutar.",
        "5": "Procedimiento perfectamente aplicable; sin ambigüedad, con canales y validaciones.",
        "_ancla": "Ancla 5: '1. Abre la App 2. Menú Servicios 3. Certificaciones...'; Ancla 1: 'gestione su certificación' sin más.",
    },
    "consistencia": {
        "_def": "¿Cumple la plantilla KCS y es coherente internamente (campos, evidencia, metadata)?",
        "1": "Estructura rota; campos contradictorios entre sí.",
        "2": "Varios campos mal usados o incoherentes con el resto.",
        "3": "Estructura presente con alguna incoherencia menor.",
        "4": "Plantilla KCS bien aplicada y coherente.",
        "5": "Plantilla impecable; evidence_pack trazable y metadata consistente con el contenido.",
        "_ancla": "Ancla 5: cause='No aplica' en howto y resolution aplicable; Ancla 1: problem habla de tarjetas y resolution de transferencias.",
    },
}


# ---------------------------------------------------------------------------
# Construcción del Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def build_eval_sheet(ws, rows: List[Dict[str, Any]]):
    content_cols = ["article_id", "title", "environment", "problem", "cause",
                    "resolution", "evidence_pack"]
    cols = content_cols + SCORE_DIMS + ["comentarios"]
    ws.append([c if c not in SCORE_DIMS else f"{c} (1-5)" for c in cols])
    _style_header(ws, len(cols))

    for r in rows:
        f = r["_fields"]
        ws.append([
            r["eval_id"], f["title"], f["environment"], f["problem"], f["cause"],
            f["resolution"], f["evidence_pack"],
            None, None, None, None, None, None,
        ])

    widths = {"article_id": 11, "title": 34, "environment": 26, "problem": 38,
              "cause": 22, "resolution": 50, "evidence_pack": 55, "comentarios": 30}
    for i, c in enumerate(cols, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(c, 13)
    # estilo de celdas + validación 1-5 en columnas de score
    score_start = len(content_cols) + 1
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                        allow_blank=True, showErrorMessage=True,
                        error="Ingresa un entero de 1 a 5", errorTitle="Valor inválido")
    ws.add_data_validation(dv)
    nrows = len(rows)
    for row in range(2, nrows + 2):
        for col in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if score_start <= col <= score_start + len(SCORE_DIMS) - 1:
                cell.fill = SCORE_FILL
        for col in range(score_start, score_start + len(SCORE_DIMS)):
            dv.add(ws.cell(row=row, column=col))
        ws.row_dimensions[row].height = 130


def build_key_sheet(ws, rows: List[Dict[str, Any]]):
    cols = ["article_id", "framework", "orig_article_id", "product_category", "interaction_ids"]
    ws.append(cols)
    _style_header(ws, len(cols))
    for r in sorted(rows, key=lambda x: x["eval_id"]):
        ws.append([
            r["eval_id"], r["framework"], r["orig_article_id"],
            r["product_category"], ", ".join(r["interaction_ids"]),
        ])
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = {"article_id": 11, "framework": 18,
            "orig_article_id": 16, "product_category": 18, "interaction_ids": 40}.get(c, 14)
    for row in range(2, len(rows) + 2):
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)


def build_rubric_sheet(ws):
    cols = ["dimensión", "definición", "nivel 1", "nivel 2", "nivel 3", "nivel 4", "nivel 5", "ejemplos ancla"]
    ws.append(cols)
    _style_header(ws, len(cols))
    for dim in SCORE_DIMS:
        r = RUBRICA[dim]
        ws.append([dim, r["_def"], r["1"], r["2"], r["3"], r["4"], r["5"], r["_ancla"]])
    widths = [16, 40, 30, 30, 30, 30, 34, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in range(2, len(SCORE_DIMS) + 2):
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).alignment = WRAP_TOP
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 150
    # nota de uso arriba
    ws.append([])
    ws.append(["Instrucciones: puntúa cada artículo de 1 (peor) a 5 (mejor) en cada dimensión "
               "de la hoja 'evaluacion'. Evaluación CIEGA: no se indica el framework generador. "
               "Usa 'comentarios' para justificar puntajes extremos (1 o 5)."])


def main() -> int:
    rng = random.Random(SEED)
    cats = load_interaction_categories()

    selected: List[Dict[str, Any]] = []
    print("Selección estratificada por product_category:")
    for fw in FRAMEWORKS:
        arts = load_framework_articles(fw, cats)
        pick = stratified_sample(arts, N_PER_FRAMEWORK, rng)
        dist = Counter(p["product_category"] for p in pick)
        print(f"  {fw:<16} {len(pick)} arts → {dict(dist)}")
        selected.extend(pick)

    control = load_framework_articles(CONTROL_FW, cats)
    cpick = stratified_sample(control, N_CONTROL, rng)
    print(f"  {CONTROL_FW:<16} {len(cpick)} arts (control) → {dict(Counter(p['product_category'] for p in cpick))}")
    selected.extend(cpick)

    # orden aleatorio + asignación de blind id
    rng.shuffle(selected)
    for idx, r in enumerate(selected, start=1):
        r["eval_id"] = f"EVAL-{idx:03d}"
        r["_fields"] = article_fields(r["article"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_eval = wb.active
    ws_eval.title = "evaluacion"
    build_eval_sheet(ws_eval, selected)
    build_key_sheet(wb.create_sheet("key"), selected)
    build_rubric_sheet(wb.create_sheet("rubrica"))
    wb.save(OUT)

    print(f"\n✅ {len(selected)} artículos → {OUT}")
    print(f"   Hojas: evaluacion (ciega, orden aleatorio) | key (mapeo) | rubrica")
    print(f"   Distribución por framework: {dict(Counter(r['framework'] for r in selected))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
