#!/usr/bin/env python
"""Análisis estadístico combinado: evaluación humana comparativa + métricas
automáticas.

Entrada:
    eval/rubrics/human_evaluation_comparative_scored.xlsx  (hojas evaluacion + key)
    eval/results/main_metrics.json                         (métricas automáticas)

Salida:
    eval/analysis/statistical_tests.json
    eval/analysis/thesis_tables.md
    eval/analysis/figures/{boxplots_dimensions,radar_comparative,heatmap_interaction_framework}.png

Pruebas: Friedman por dimensión (bloques = interacción), W de Kendall (tamaño de
efecto), post-hoc Nemenyi (CD vía studentized range) + Wilcoxon pareado con
Bonferroni. Correlación humano↔automático (Spearman).
"""
from __future__ import annotations

import json
import re
import statistics
from itertools import combinations
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
from scipy import stats
from openpyxl import load_workbook

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCORED = PROJECT_ROOT / "eval/rubrics/human_evaluation_comparative_scored.xlsx"
AUTO = PROJECT_ROOT / "eval/results/main_metrics.json"
OUTDIR = PROJECT_ROOT / "eval/analysis"
FIGDIR = OUTDIR / "figures"

DIMS = ["claridad", "exactitud", "completitud", "accionabilidad", "consistencia"]
FRAMEWORKS = ["langgraph", "crewai", "openai_agents", "baseline_prompt"]
THREE_FW = ["langgraph", "crewai", "openai_agents"]  # sin baseline (H2/H3)
FW_SHORT = {"langgraph": "LangGraph", "crewai": "CrewAI",
            "openai_agents": "OpenAI", "baseline_prompt": "Baseline"}
ALPHA = 0.05


# ---------------------------------------------------------------------------
# 1) Parseo del Excel scored
# ---------------------------------------------------------------------------
def parse_scores() -> List[Dict[str, Any]]:
    wb = load_workbook(SCORED, data_only=True)
    ev = wb["evaluacion"]
    key = wb["key"]
    # mapeo (interaction_id, version) -> framework
    fw_of: Dict[Tuple[str, str], str] = {}
    cat_of: Dict[str, str] = {}
    for r in range(2, key.max_row + 1):
        iid, ver, fw, _aid, cat = [key.cell(r, c).value for c in range(1, 6)]
        if iid and ver:
            fw_of[(iid, str(ver).strip())] = fw
            cat_of[iid] = cat

    records: List[Dict[str, Any]] = []
    current_iid = None
    for r in range(2, ev.max_row + 1):
        c1 = ev.cell(r, 1).value
        if c1 is None:
            continue
        s = str(c1).strip()
        m = re.match(r"(INT-\d{4}-\d+)", s)
        if m:
            current_iid = m.group(1)
            continue
        if s in ("A", "B", "C", "D") and current_iid:
            scores = [ev.cell(r, 5 + i).value for i in range(len(DIMS))]
            fw = fw_of.get((current_iid, s))
            if fw is None:
                raise ValueError(f"Sin framework para {current_iid} versión {s}")
            if any(v is None for v in scores):
                raise ValueError(f"Puntuación incompleta en {current_iid} versión {s}: {scores}")
            rec = {"interaction_id": current_iid, "version": s, "framework": fw,
                   "product_category": cat_of.get(current_iid)}
            for d, v in zip(DIMS, scores):
                rec[d] = float(v)
            records.append(rec)
    return records


# ---------------------------------------------------------------------------
# 2) Estadística descriptiva
# ---------------------------------------------------------------------------
def descriptive(records) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for fw in FRAMEWORKS:
        out[fw] = {}
        for d in DIMS:
            vals = [r[d] for r in records if r["framework"] == fw]
            out[fw][d] = {
                "n": len(vals),
                "median": statistics.median(vals),
                "mean": round(statistics.mean(vals), 3),
                "std": round(statistics.pstdev(vals), 3),
                "min": min(vals),
                "max": max(vals),
            }
        allvals = [r[d] for r in records if r["framework"] == fw for d in DIMS]
        out[fw]["_overall"] = {
            "mean": round(statistics.mean(allvals), 3),
            "median": statistics.median(allvals),
            "std": round(statistics.pstdev(allvals), 3),
        }
    return out


# ---------------------------------------------------------------------------
# 3) Matriz bloque×framework por dimensión + Friedman + Kendall W + post-hoc
# ---------------------------------------------------------------------------
def _matrix(records, dim) -> Tuple[List[str], np.ndarray]:
    iids = sorted({r["interaction_id"] for r in records})
    M = np.full((len(iids), len(FRAMEWORKS)), np.nan)
    idx = {iid: i for i, iid in enumerate(iids)}
    fidx = {fw: j for j, fw in enumerate(FRAMEWORKS)}
    for r in records:
        M[idx[r["interaction_id"]], fidx[r["framework"]]] = r[dim]
    return iids, M


def nemenyi_cd(k: int, n: int, alpha: float = ALPHA) -> float:
    q = stats.studentized_range.ppf(1 - alpha, k, np.inf) / np.sqrt(2)
    return float(q * np.sqrt(k * (k + 1) / (6.0 * n)))


def friedman_block(records) -> Dict[str, Any]:
    res: Dict[str, Any] = {}
    for d in DIMS:
        iids, M = _matrix(records, d)
        cols = [M[:, j] for j in range(len(FRAMEWORKS))]
        chi2, p = stats.friedmanchisquare(*cols)
        n, k = M.shape
        W = float(chi2 / (n * (k - 1)))  # Kendall's W
        # rangos promedio (1 = mejor): rank descendente por bloque
        ranks = np.zeros_like(M)
        for i in range(n):
            ranks[i] = stats.rankdata(-M[i])  # mayor score -> rank menor (mejor)
        avg_ranks = {FRAMEWORKS[j]: float(ranks[:, j].mean()) for j in range(k)}
        entry = {
            "friedman_chi2": round(float(chi2), 4),
            "p_value": float(p),
            "significant": bool(p < ALPHA),
            "kendall_w": round(W, 4),
            "kendall_w_interpretacion": _w_interp(W),
            "avg_ranks": {f: round(v, 3) for f, v in avg_ranks.items()},
            "n_blocks": int(n),
            "k_treatments": int(k),
        }
        if p < ALPHA:
            cd = nemenyi_cd(k, n)
            nem = {}
            for a, b in combinations(FRAMEWORKS, 2):
                diff = abs(avg_ranks[a] - avg_ranks[b])
                nem[f"{a}__vs__{b}"] = {
                    "rank_diff": round(diff, 3), "CD": round(cd, 3),
                    "significant": bool(diff > cd),
                }
            entry["nemenyi"] = {"critical_difference": round(cd, 3), "pairs": nem}
            # Wilcoxon pareado + Bonferroni
            pairs = list(combinations(range(k), 2))
            wil = {}
            for ia, ib in pairs:
                a, b = M[:, ia], M[:, ib]
                if np.allclose(a, b):
                    stat, pw = float("nan"), 1.0
                else:
                    try:
                        stat, pw = stats.wilcoxon(a, b, zero_method="wilcox")
                    except ValueError:
                        stat, pw = float("nan"), 1.0
                wil[f"{FRAMEWORKS[ia]}__vs__{FRAMEWORKS[ib]}"] = {
                    "stat": (None if np.isnan(stat) else round(float(stat), 3)),
                    "p_raw": round(float(pw), 5),
                    "p_bonferroni": round(min(1.0, float(pw) * len(pairs)), 5),
                    "significant_bonf": bool(pw * len(pairs) < ALPHA),
                }
            entry["wilcoxon_bonferroni"] = wil
        res[d] = entry
    return res


def _w_interp(w: float) -> str:
    if w < 0.1: return "acuerdo muy débil"
    if w < 0.3: return "acuerdo débil"
    if w < 0.5: return "acuerdo moderado"
    if w < 0.7: return "acuerdo fuerte"
    return "acuerdo muy fuerte"


# ---------------------------------------------------------------------------
# 4) Combinar con métricas automáticas + correlación
# ---------------------------------------------------------------------------
def auto_metrics() -> Dict[str, Dict[str, float]]:
    rep = json.loads(AUTO.read_text(encoding="utf-8"))["frameworks"]
    out = {}
    for fw in FRAMEWORKS:
        e = rep.get(fw, {})
        q = e.get("quality", {}) or {}
        eng = e.get("engineering", {}) or {}
        cons = e.get("consolidation", {}) or {}
        sim = q.get("reference_similarity_max_avg")
        out[fw] = {
            "simK": round((sim or 0) * 100, 2) if sim is not None else None,
            "kcs_pct": round((q.get("kcs_compliance_avg") or 0) * 100, 1),
            "evidence_pct": round((q.get("evidence_coverage_avg") or 0) * 100, 1),
            "coverage_pct": round(cons.get("interaction_coverage_pct_avg") or 0, 1),
            "cost_usd": round((eng.get("cost_usd") or {}).get("median") or 0, 3),
            "tool_calls": round(eng.get("tool_calls_avg") or 0, 1),
            "latency_s": round((eng.get("latency_seconds") or {}).get("median") or 0, 1),
            "loc": eng.get("loc"),
        }
    return out


def correlations(desc, autom) -> Dict[str, Any]:
    human = [desc[fw]["_overall"]["mean"] for fw in FRAMEWORKS]
    out = {}
    for metric in ["simK", "cost_usd", "tool_calls", "coverage_pct"]:
        auto_vals = [autom[fw][metric] for fw in FRAMEWORKS]
        if any(v is None for v in auto_vals):
            continue
        rho, p = stats.spearmanr(human, auto_vals)
        rp, pp = stats.pearsonr(human, auto_vals)
        out[metric] = {
            "spearman_rho": round(float(rho), 3), "spearman_p": round(float(p), 4),
            "pearson_r": round(float(rp), 3), "pearson_p": round(float(pp), 4),
            "human_means": dict(zip(FRAMEWORKS, human)),
            "auto_values": dict(zip(FRAMEWORKS, auto_vals)),
            "n": len(FRAMEWORKS),
        }
    return out


# ---------------------------------------------------------------------------
# 5) Figuras
# ---------------------------------------------------------------------------
def fig_boxplots(records):
    fig, axes = plt.subplots(1, len(DIMS), figsize=(20, 4.5), sharey=True)
    for ax, d in zip(axes, DIMS):
        data = [[r[d] for r in records if r["framework"] == fw] for fw in FRAMEWORKS]
        bp = ax.boxplot(data, patch_artist=True, widths=0.6)
        for patch, color in zip(bp["boxes"], ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]):
            patch.set_facecolor(color); patch.set_alpha(0.7)
        ax.set_title(d, fontsize=12, fontweight="bold")
        ax.set_xticks(range(1, len(FRAMEWORKS) + 1))
        ax.set_xticklabels([FW_SHORT[f] for f in FRAMEWORKS], rotation=35, ha="right", fontsize=9)
        ax.set_ylim(0.5, 5.5); ax.grid(axis="y", alpha=0.3)
    axes[0].set_ylabel("Puntuación (1-5)", fontsize=11)
    fig.suptitle("Distribución de puntuaciones humanas por dimensión y framework", fontsize=14, fontweight="bold")
    fig.tight_layout()
    fig.savefig(FIGDIR / "boxplots_dimensions.png", dpi=150, bbox_inches="tight")
    plt.close(fig)


def fig_radar(desc, autom):
    # 8 ejes: 5 humanas (mean) + 3 automáticas clave. Normalización min-max por eje
    # (dirección: mayor=mejor; cost y tool_calls invertidos).
    axes_labels = DIMS + ["simK", "efic_costo", "efic_tools"]
    raw = {}
    for fw in FRAMEWORKS:
        vals = [desc[fw][d]["mean"] for d in DIMS]
        vals += [autom[fw]["simK"], -autom[fw]["cost_usd"], -autom[fw]["tool_calls"]]
        raw[fw] = vals
    arr = np.array([raw[fw] for fw in FRAMEWORKS], dtype=float)
    mn, mx = arr.min(axis=0), arr.max(axis=0)
    rng = np.where(mx - mn == 0, 1, mx - mn)
    norm = (arr - mn) / rng  # 0..1 por eje
    N = len(axes_labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]
    for i, fw in enumerate(FRAMEWORKS):
        v = norm[i].tolist(); v += v[:1]
        ax.plot(angles, v, color=colors[i], linewidth=2, label=FW_SHORT[fw])
        ax.fill(angles, v, color=colors[i], alpha=0.12)
    ax.set_xticks(angles[:-1]); ax.set_xticklabels(axes_labels, fontsize=10)
    ax.set_yticks([0.25, 0.5, 0.75, 1.0]); ax.set_ylim(0, 1)
    ax.set_title("Radar comparativo (normalizado min-max por eje;\nmayor = mejor; costo/tools invertidos)",
                 fontsize=12, fontweight="bold", pad=24)
    ax.legend(loc="upper right", bbox_to_anchor=(1.25, 1.1))
    fig.savefig(FIGDIR / "radar_comparative.png", dpi=150, bbox_inches="tight")
    plt.close(fig)


def fig_heatmap(records):
    iids = sorted({r["interaction_id"] for r in records})
    M = np.full((len(iids), len(FRAMEWORKS)), np.nan)
    idx = {iid: i for i, iid in enumerate(iids)}
    fidx = {fw: j for j, fw in enumerate(FRAMEWORKS)}
    agg: Dict[Tuple[int, int], List[float]] = {}
    for r in records:
        agg.setdefault((idx[r["interaction_id"]], fidx[r["framework"]]), []).extend(r[d] for d in DIMS)
    for (i, j), vals in agg.items():
        M[i, j] = np.mean(vals)
    fig, ax = plt.subplots(figsize=(7, 11))
    im = ax.imshow(M, cmap="RdYlGn", vmin=1, vmax=5, aspect="auto")
    ax.set_xticks(range(len(FRAMEWORKS))); ax.set_xticklabels([FW_SHORT[f] for f in FRAMEWORKS], fontsize=10)
    ax.set_yticks(range(len(iids))); ax.set_yticklabels(iids, fontsize=8)
    for i in range(len(iids)):
        for j in range(len(FRAMEWORKS)):
            if not np.isnan(M[i, j]):
                ax.text(j, i, f"{M[i,j]:.1f}", ha="center", va="center", fontsize=8,
                        color="black")
    ax.set_title("Puntuación media (5 dim.) por interacción × framework", fontsize=12, fontweight="bold")
    fig.colorbar(im, ax=ax, shrink=0.5, label="media 1-5")
    fig.tight_layout()
    fig.savefig(FIGDIR / "heatmap_interaction_framework.png", dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# 6) Hipótesis H1–H4
# ---------------------------------------------------------------------------
def _load_summary(fw) -> Dict[str, Any]:
    return json.loads((PROJECT_ROOT / "runs/experiment" / fw / "experiment_summary.json").read_text(encoding="utf-8"))


def failure_rates() -> Dict[str, Any]:
    """H1: tasa de fallos = errores / lotes totales (n_runs × interacciones)."""
    out = {}
    for fw in THREE_FW:
        s = _load_summary(fw)
        batches = s["n_runs"] * s["interaction_count"]
        errs = s["totals"]["errors"]
        out[fw] = {
            "errors_total": errs,
            "batches_total": batches,
            "errors_per_run": [r["errors"] for r in s["runs"]],
            "failure_rate_pct": round(errs / batches * 100, 2),
        }
    return out


def friedman_subset(records, frameworks, dim) -> Dict[str, Any]:
    """Friedman restringido a un subconjunto de generadores (H2: sin baseline)."""
    iids = sorted({r["interaction_id"] for r in records})
    fidx = {fw: j for j, fw in enumerate(frameworks)}
    M = np.full((len(iids), len(frameworks)), np.nan)
    idx = {iid: i for i, iid in enumerate(iids)}
    for r in records:
        if r["framework"] in fidx:
            M[idx[r["interaction_id"]], fidx[r["framework"]]] = r[dim]
    cols = [M[:, j] for j in range(len(frameworks))]
    chi2, p = stats.friedmanchisquare(*cols)
    n, k = M.shape
    W = float(chi2 / (n * (k - 1)))
    means = {fw: round(float(np.nanmean(M[:, fidx[fw]])), 3) for fw in frameworks}
    out = {"friedman_chi2": round(float(chi2), 4), "p_value": float(p),
           "significant": bool(p < ALPHA), "kendall_w": round(W, 4),
           "means": means, "best": max(means, key=means.get)}
    if p < ALPHA:
        pairs = list(combinations(range(k), 2))
        wil = {}
        for ia, ib in pairs:
            a, b = M[:, ia], M[:, ib]
            try:
                stat, pw = stats.wilcoxon(a, b, zero_method="wilcox") if not np.allclose(a, b) else (float("nan"), 1.0)
            except ValueError:
                stat, pw = float("nan"), 1.0
            wil[f"{frameworks[ia]}__vs__{frameworks[ib]}"] = {
                "p_raw": round(float(pw), 5),
                "p_bonferroni": round(min(1.0, float(pw) * len(pairs)), 5),
                "significant_bonf": bool(pw * len(pairs) < ALPHA)}
        out["wilcoxon_bonferroni"] = wil
    return out


def inter_run_cv() -> Dict[str, Any]:
    """H3: coeficiente de variación (CV=SD/media) entre los 3 runs por métrica."""
    out = {}
    for fw in THREE_FW:
        s = _load_summary(fw)
        runs = s["runs"]
        metrics = {
            "cost_usd": [r["cost_usd"] for r in runs],
            "tool_calls": [r["tool_calls"] for r in runs],
            "articles": [r["articles_generated"] for r in runs],
            "errors": [r["errors"] for r in runs],
        }
        cv = {}
        for m, vals in metrics.items():
            mean = statistics.mean(vals)
            sd = statistics.pstdev(vals)
            cv[m] = {"values": vals, "mean": round(mean, 3), "sd": round(sd, 3),
                     "cv_pct": (round(sd / mean * 100, 2) if mean else None)}
        # CV promedio SOLO sobre métricas sustantivas (costo, tools, artículos).
        # Se excluye `errors`: es un conteo entero diminuto cuyo CV relativo es
        # enorme y ruidoso (p.ej. [1,0,1]→CV~71%), dominaría el promedio y
        # distorsionaría la reproducibilidad. La variabilidad de errores ya se
        # cubre en H1 (tasa de fallos).
        substantive = ["cost_usd", "tool_calls", "articles"]
        valid = [cv[m]["cv_pct"] for m in substantive if cv[m]["cv_pct"] is not None]
        out[fw] = {"per_metric": cv, "mean_cv_pct": round(statistics.mean(valid), 2) if valid else None}
    return out


def evaluate_hypotheses(records, desc) -> Dict[str, Any]:
    fr = failure_rates()
    # H1: LangGraph menor tasa de fallos
    ranked = sorted(THREE_FW, key=lambda f: fr[f]["failure_rate_pct"])
    lg_lowest = ranked[0] == "langgraph"
    h1 = {
        "statement": "H1: LangGraph tendrá menor tasa de fallos que los otros frameworks.",
        "failure_rates_pct": {f: fr[f]["failure_rate_pct"] for f in THREE_FW},
        "ranking_menor_a_mayor": ranked,
        "verdict": "REFUTADA" if not lg_lowest else "SOPORTADA",
        "detail": (f"LangGraph tiene la tasa de fallo MÁS ALTA "
                   f"({fr['langgraph']['failure_rate_pct']}%); la menor es "
                   f"{FW_SHORT[ranked[0]]} ({fr[ranked[0]]['failure_rate_pct']}%).")
                  if not lg_lowest else
                  f"LangGraph tiene la menor tasa de fallo ({fr['langgraph']['failure_rate_pct']}%).",
        "_raw": fr,
    }
    # H2: CrewAI mayor calidad en claridad y accionabilidad (3 frameworks)
    h2_dims = {}
    direction_ok = True
    for d in ["claridad", "accionabilidad"]:
        sub = friedman_subset(records, THREE_FW, d)
        crewai_best = sub["best"] == "crewai"
        direction_ok = direction_ok and crewai_best
        h2_dims[d] = {**sub, "crewai_is_best": crewai_best}
    any_sig = any(h2_dims[d]["significant"] for d in h2_dims)
    h2 = {
        "statement": "H2: CrewAI tendrá mayor calidad percibida en claridad y accionabilidad.",
        "by_dimension": h2_dims,
        "verdict": ("SOPORTADA" if direction_ok and any_sig else
                    "PARCIALMENTE SOPORTADA" if direction_ok else "REFUTADA"),
        "detail": ("CrewAI obtiene la media más alta entre los 3 frameworks en ambas "
                   "dimensiones, pero el Friedman restringido a los 3 frameworks "
                   f"{'SÍ' if any_sig else 'NO'} alcanza significancia (las diferencias "
                   "entre frameworks no son estadísticamente concluyentes).")
                  if direction_ok else
                  "CrewAI no es el mejor en ambas dimensiones entre los 3 frameworks.",
    }
    # H3: reproducibilidad (CV inter-run)
    cv = inter_run_cv()
    most_repro = min(THREE_FW, key=lambda f: cv[f]["mean_cv_pct"])
    h3 = {
        "statement": "H3: Mejor instrumentación → mejor reproducibilidad (menor CV inter-run).",
        "inter_run_cv": cv,
        "most_reproducible": most_repro,
        "verdict": "EVALUADA (parcial)",
        "detail": (f"{FW_SHORT[most_repro]} es el más reproducible "
                   f"(CV medio {cv[most_repro]['mean_cv_pct']}%). "
                   "Nota: la 'instrumentación' no se midió como variable independiente; "
                   "se reporta la reproducibilidad (CV) como evidencia observacional, no causal."),
    }
    # H4: evidence pack → exactitud (ablación no ejecutada)
    h4 = {
        "statement": "H4: La presencia de evidence pack mejora la exactitud.",
        "verdict": "NO EVALUADA",
        "detail": ("LIMITACIÓN: requiere la ablación `no_evidence` (generar artículos sin "
                   "evidence pack y comparar exactitud), que NO se ejecutó en esta batería. "
                   "El flag `--ablation no_evidence` existe pero sólo se registra en metadata; "
                   "el wiring en los runners está pendiente. No se puede contrastar con los datos actuales."),
    }
    return {"H1": h1, "H2": h2, "H3": h3, "H4": h4}


# ---------------------------------------------------------------------------
# Tablas markdown
# ---------------------------------------------------------------------------
def write_tables(desc, friedman, autom, corr, hyp, n_records):
    L = []
    L.append("# Tablas de tesis — Análisis combinado (humano + automático)\n")
    L.append(f"_n = {n_records} evaluaciones ({n_records//len(FRAMEWORKS)} interacciones × {len(FRAMEWORKS)} generadores), "
             "evaluación humana comparativa ciega._\n")

    L.append("## 1. Estadística descriptiva — puntuación humana (media ± std, [mediana])\n")
    head = "| framework | " + " | ".join(DIMS) + " | **global** |"
    L.append(head); L.append("|" + "---|" * (len(DIMS) + 2))
    for fw in FRAMEWORKS:
        cells = []
        for d in DIMS:
            s = desc[fw][d]
            cells.append(f"{s['mean']:.2f}±{s['std']:.2f} [{s['median']:.0f}]")
        ov = desc[fw]["_overall"]
        L.append(f"| {FW_SHORT[fw]} | " + " | ".join(cells) + f" | **{ov['mean']:.2f}±{ov['std']:.2f}** |")
    L.append("")

    L.append("## 2. Pruebas de Friedman por dimensión (bloques = interacción)\n")
    L.append("| dimensión | χ² | p | sig. | W Kendall | interpretación | mejor (rango↓) |")
    L.append("|---|---|---|---|---|---|---|")
    for d in DIMS:
        f = friedman[d]
        best = min(f["avg_ranks"], key=f["avg_ranks"].get)
        L.append(f"| {d} | {f['friedman_chi2']:.2f} | {f['p_value']:.4f} | "
                 f"{'✅' if f['significant'] else '—'} | {f['kendall_w']:.3f} | "
                 f"{f['kendall_w_interpretacion']} | {FW_SHORT[best]} ({f['avg_ranks'][best]:.2f}) |")
    L.append("")

    # post-hoc para dimensiones significativas
    sig_dims = [d for d in DIMS if friedman[d]["significant"]]
    if sig_dims:
        L.append("### 2b. Post-hoc (dimensiones significativas)\n")
        for d in sig_dims:
            f = friedman[d]
            L.append(f"**{d}** — Nemenyi CD={f['nemenyi']['critical_difference']:.3f}:")
            for pair, v in f["nemenyi"]["pairs"].items():
                if v["significant"]:
                    a, b = pair.split("__vs__")
                    L.append(f"  - {FW_SHORT[a]} vs {FW_SHORT[b]}: Δrango={v['rank_diff']:.2f} > CD → **significativo**")
            sigp = [p for p, v in f["nemenyi"]["pairs"].items() if v["significant"]]
            if not sigp:
                L.append("  - (ningún par supera la CD de Nemenyi pese a Friedman significativo)")
            L.append("")
    else:
        L.append("_Ninguna dimensión alcanzó significancia en Friedman (p<0.05); sin post-hoc._\n")

    L.append("## 3. Tabla consolidada — calidad humana + eficiencia automática\n")
    L.append("| framework | humano global | simK | KCS% | evid% | cobertura% | costo$ | tool_calls | LOC |")
    L.append("|---|---|---|---|---|---|---|---|---|")
    for fw in FRAMEWORKS:
        a = autom[fw]; ov = desc[fw]["_overall"]["mean"]
        L.append(f"| {FW_SHORT[fw]} | {ov:.2f} | {a['simK']} | {a['kcs_pct']} | {a['evidence_pct']} | "
                 f"{a['coverage_pct']} | {a['cost_usd']} | {a['tool_calls']} | {a['loc']} |")
    L.append("")

    L.append("## 4. Correlación humano ↔ automático (n=4 generadores)\n")
    L.append("| métrica auto | Spearman ρ | p | Pearson r | p | lectura |")
    L.append("|---|---|---|---|---|---|")
    for m, v in corr.items():
        sign = "↑" if v["spearman_rho"] > 0 else "↓"
        L.append(f"| {m} | {v['spearman_rho']:.3f} | {v['spearman_p']:.3f} | "
                 f"{v['pearson_r']:.3f} | {v['pearson_p']:.3f} | {sign} |")
    L.append("\n> ⚠️ n=4 generadores: las correlaciones son exploratorias, sin potencia para significancia. "
             "Interpretar como tendencia, no como prueba.\n")

    L.append("## 5. Figuras\n")
    L.append("- `figures/boxplots_dimensions.png` — distribución por dimensión × framework")
    L.append("- `figures/radar_comparative.png` — 5 dim. humanas + 3 métricas automáticas (normalizado)")
    L.append("- `figures/heatmap_interaction_framework.png` — media por interacción × framework\n")

    L.append("## 6. Contraste de hipótesis H1–H4\n")
    icon = {"SOPORTADA": "✅", "PARCIALMENTE SOPORTADA": "🟡", "REFUTADA": "❌",
            "NO EVALUADA": "⚪", "EVALUADA (parcial)": "🟡"}

    h1 = hyp["H1"]
    L.append(f"### {icon[h1['verdict']]} H1 — {h1['verdict']}")
    L.append(f"*{h1['statement']}*\n")
    L.append("| framework | tasa de fallo (errores/lotes) |")
    L.append("|---|---|")
    for f in THREE_FW:
        L.append(f"| {FW_SHORT[f]} | {h1['failure_rates_pct'][f]}% |")
    L.append(f"\n{h1['detail']}\n")

    h2 = hyp["H2"]
    L.append(f"### {icon[h2['verdict']]} H2 — {h2['verdict']}")
    L.append(f"*{h2['statement']}*\n")
    L.append("| dimensión | medias (LG / CW / OA) | mejor 3-fw | Friedman p (3-fw) | sig. |")
    L.append("|---|---|---|---|---|")
    for d in ["claridad", "accionabilidad"]:
        s = h2["by_dimension"][d]
        m = s["means"]
        L.append(f"| {d} | {m['langgraph']:.2f} / {m['crewai']:.2f} / {m['openai_agents']:.2f} | "
                 f"{FW_SHORT[s['best']]} | {s['p_value']:.4f} | {'✅' if s['significant'] else '—'} |")
    L.append(f"\n{h2['detail']}\n")

    h3 = hyp["H3"]
    L.append(f"### {icon[h3['verdict']]} H3 — {h3['verdict']}")
    L.append(f"*{h3['statement']}*\n")
    L.append("| framework | CV costo | CV tool_calls | CV artículos | CV medio |")
    L.append("|---|---|---|---|---|")
    for f in THREE_FW:
        c = h3["inter_run_cv"][f]["per_metric"]
        L.append(f"| {FW_SHORT[f]} | {c['cost_usd']['cv_pct']}% | {c['tool_calls']['cv_pct']}% | "
                 f"{c['articles']['cv_pct']}% | **{h3['inter_run_cv'][f]['mean_cv_pct']}%** |")
    L.append(f"\n{h3['detail']}\n")

    h4 = hyp["H4"]
    L.append(f"### {icon[h4['verdict']]} H4 — {h4['verdict']}")
    L.append(f"*{h4['statement']}*\n")
    L.append(f"{h4['detail']}\n")

    (OUTDIR / "thesis_tables.md").write_text("\n".join(L), encoding="utf-8")


def main() -> int:
    FIGDIR.mkdir(parents=True, exist_ok=True)
    records = parse_scores()
    assert len(records) == 64, f"esperaba 64 registros, obtuve {len(records)}"
    desc = descriptive(records)
    friedman = friedman_block(records)
    autom = auto_metrics()
    corr = correlations(desc, autom)
    hyp = evaluate_hypotheses(records, desc)

    fig_boxplots(records)
    fig_radar(desc, autom)
    fig_heatmap(records)

    payload = {
        "n_records": len(records),
        "n_interactions": len({r["interaction_id"] for r in records}),
        "frameworks": FRAMEWORKS,
        "dimensions": DIMS,
        "alpha": ALPHA,
        "descriptive": desc,
        "friedman_by_dimension": friedman,
        "automatic_metrics": autom,
        "correlations_human_vs_auto": corr,
        "hypotheses": hyp,
    }
    (OUTDIR).mkdir(parents=True, exist_ok=True)
    (OUTDIR / "statistical_tests.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_tables(desc, friedman, autom, corr, hyp, len(records))

    # resumen a consola
    print(f"✅ {len(records)} registros · {payload['n_interactions']} interacciones")
    print("\nFriedman por dimensión:")
    for d in DIMS:
        f = friedman[d]
        print(f"  {d:<14} χ²={f['friedman_chi2']:6.2f}  p={f['p_value']:.4f}  "
              f"{'SIG' if f['significant'] else '  -'}  W={f['kendall_w']:.3f}")
    print("\nMedia humana global por framework:")
    for fw in FRAMEWORKS:
        print(f"  {FW_SHORT[fw]:<10} {desc[fw]['_overall']['mean']:.3f}")
    print(f"\nSalida en {OUTDIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
